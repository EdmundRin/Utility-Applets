import openpyxl
import copy
from openpyxl import Workbook
from tkinter import filedialog, Tk

def select_file():
    root = Tk()
    root.withdraw()
    return filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel files", "*.xlsx")])

def select_save_path():
    root = Tk()
    root.withdraw()
    return filedialog.asksaveasfilename(title="保存Excel文件", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

def choose_sheet(wb):
    sheets = wb.sheetnames
    print("\n可用工作表: ")
    for idx, name in enumerate(sheets):
        print(f"{idx + 1}: {name}")
    while True:
        try:
            choice = int(input("请输入要选择的工作表编号: "))
            if 1 <= choice <= len(sheets):
                return wb[sheets[choice - 1]]
            else:
                print("编号超出范围，请重新输入。")
        except ValueError:
            print("请输入有效的数字编号。")

def display_column_headers(sheet):
    print("\n首行(通常为列名): ")
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=col).value
        print(f"列 {col}: {value}")

def display_row_headers(sheet):
    print("\n首列(通常为行名): ")
    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row=row, column=1).value
        print(f"行 {row}: {value}")

def find_rows_with_red_font(sheet, target_col):
    red_rows = []
    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=target_col)
        font_color = cell.font.color

        if font_color is None:
            continue

        rgb = None
        # If color is RGB type (like 'FFFF0000')
        if hasattr(font_color, 'rgb') and isinstance(font_color.rgb, str):
            rgb = font_color.rgb.upper()
        # If it's a Color object with type='rgb' and value is string
        elif getattr(font_color, 'type', None) == 'rgb' and isinstance(font_color.value, str):
            rgb = font_color.value.upper()

        if rgb in ("FF0000", "FFFF0000"):  # Red in common formats
            red_rows.append(row)

    return red_rows

def copy_selected_cells(sheet, selected_rows, selected_cols, keep_style=False):
    new_wb = Workbook()
    new_sheet = new_wb.active

    for new_i, row in enumerate(selected_rows, start=1):
        for new_j, col in enumerate(selected_cols, start=1):
            src_cell = sheet.cell(row=row, column=col)
            dst_cell = new_sheet.cell(row=new_i, column=new_j, value=src_cell.value)

            if keep_style:
                dst_cell.font = copy.copy(src_cell.font)
                dst_cell.fill = copy.copy(src_cell.fill)
                dst_cell.alignment = copy.copy(src_cell.alignment)
                dst_cell.border = copy.copy(src_cell.border)
                dst_cell.number_format = src_cell.number_format

    save_path = select_save_path()
    if save_path:
        new_wb.save(save_path)
        print(f"\n已保存到: {save_path}")
    else:
        print("未选择保存路径，操作已取消。")

def copy_rows_or_cols(sheet, indices, mode, keep_style=False):
    new_wb = Workbook()
    new_sheet = new_wb.active

    if mode == "row":
        for new_i, row_num in enumerate(indices, start=1):
            for j, src_cell in enumerate(sheet[row_num], start=1):
                dst_cell = new_sheet.cell(row=new_i, column=j, value=src_cell.value)
                if keep_style:
                    dst_cell.font = copy.copy(src_cell.font)
                    dst_cell.fill = copy.copy(src_cell.fill)
                    dst_cell.alignment = copy.copy(src_cell.alignment)
                    dst_cell.border = copy.copy(src_cell.border)
                    dst_cell.number_format = src_cell.number_format
    elif mode == "col":
        for new_j, col_num in enumerate(indices, start=1):
            for i in range(1, sheet.max_row + 1):
                src_cell = sheet.cell(row=i, column=col_num)
                dst_cell = new_sheet.cell(row=i, column=new_j, value=src_cell.value)
                if keep_style:
                    dst_cell.font = copy.copy(src_cell.font)
                    dst_cell.fill = copy.copy(src_cell.fill)
                    dst_cell.alignment = copy.copy(src_cell.alignment)
                    dst_cell.border = copy.copy(src_cell.border)
                    dst_cell.number_format = src_cell.number_format

    save_path = select_save_path()
    if save_path:
        new_wb.save(save_path)
        print(f"\n已保存到: {save_path}")
    else:
        print("未选择保存路径，操作已取消。")

def main():
    file_path = select_file()
    if not file_path:
        print("未选择文件，程序已退出。")
        return

    wb = openpyxl.load_workbook(file_path)
    sheet = choose_sheet(wb)

    mode = input("\n请输入模式 (row, col, more): ").strip().lower()

    if mode == "row" or mode == "col":
        try:
            index_str = input(f"请输入要复制的{mode}编号(用英文逗号分隔，如 3,1,5): ")
            indices = [int(x.strip()) for x in index_str.split(",")]

            keep_style = input("是否保留单元格格式? (y/n): ").strip().lower() == "y"

            copy_rows_or_cols(sheet, indices, mode, keep_style)

        except Exception as e:
            print(f"发生错误: {e}")
        return

    elif mode == "more":
        try:
            show_col = input("是否查看首行列名？(y/n): ").strip().lower()
            if show_col == "y":
                display_column_headers(sheet)

            filter_col = int(input("请输入用于筛选红色字体的列编号: "))
            red_rows = find_rows_with_red_font(sheet, filter_col)

            if not red_rows:
                print("未找到任何字体为红色的单元格。")
                return

            print(f"共找到 {len(red_rows)} 行字体为红色的单元格。")

            show_copy_col = input("是否查看首行列名以辅助选择提取列？(y/n): ").strip().lower()
            if show_copy_col == "y":
                display_column_headers(sheet)

            col_input = input("请输入要提取的列编号 (按顺序，用英文逗号分隔，如 4,1): ")
            selected_cols = [int(c.strip()) for c in col_input.split(",") if c.strip().isdigit()]
            if not selected_cols:
                print("未输入有效的列编号。")
                return

            keep_style = input("是否保留单元格格式? (y/n): ").strip().lower() == "y"

            copy_selected_cells(sheet, red_rows, selected_cols, keep_style)

        except Exception as e:
            print(f"发生错误: {e}")
        return

    else:
        print("无效的模式输入，程序结束。")

if __name__ == "__main__":
    main()